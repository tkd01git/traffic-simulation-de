import pandas as pd
import openpyxl
import re
import os

def column_to_letter(column_index):
    """Excelの列番号をアルファベットに変換"""
    letter = ""
    while column_index >= 0:
        letter = chr(column_index % 26 + ord('A')) + letter
        column_index = column_index // 26 - 1
    return letter

file_path = "C://Users//YuheiTakada//OneDrive//デスクトップ//traffic-simulation-de//prodata800-15.xlsx"

# 最終編集先Excelファイル
output_file = "λrankresult15.xlsx"

# prodata.xlsxを読み込む
workbook = openpyxl.load_workbook(file_path)
sheet = workbook["Sheet1"]  # Sheet1を対象とする

detector_col_index = None

# ヘッダー行（最初の行）をチェックして'列'のインデックスを取得
for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == "detector1":
        detector_col_index = col[0].column - 1  # 0ベースに変換
        break

if detector_col_index is not None:
    # 列番号をアルファベットに変換
    detector_col_letter = column_to_letter(detector_col_index)
    # 2つ前の列番号を計算
    if detector_col_index >= 2:
        two_columns_before_letter = column_to_letter(detector_col_index - 2)
    else:
        print("2つ前の列は存在しません。")
else:
    print("Detector1が見つかりません。")
import pandas as pd
import numpy as np
from scipy.linalg import eigh


start_col = 'A'
end_col = two_columns_before_letter
sheet_name1 = 'Sheet1'
sheet_name2 = 'Sheet2'

# Sheet1とSheet2のデータの取得
df1 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None)
df2 = pd.read_excel(file_path, sheet_name=sheet_name2, usecols=f'{start_col}:{end_col}', header=None)

# 新しいデータフレームの作成
df3 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=1, nrows=20)  # 2~31行目
df4 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=4, nrows=20)  # 5~34行目
df5 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=7, nrows=20)  # 8~37行目
df6 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=10, nrows=20) # 11~40行目
df7 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=13, nrows=20)  # 5~34行目
df8 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=16, nrows=20)  # 5~34行目
df9 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=19, nrows=20)  # 5~34行目
df10 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=22, nrows=20)  # 5~34行目


time_data = df1.iloc[0].values.flatten()
speed_data = df2.iloc[44].values.flatten() 
dens_data = df2.iloc[99].values.flatten()

# Laplacianの作成
n = 20
A = np.zeros((n, n))
for i in range(n - 1):
    A[i, i + 1] = 1
    A[i + 1, i] = 1
L = np.diag(np.sum(A, axis=1)) - A
L_normalized = np.dot(np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1)))), 
                      np.dot(L, np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1))))))



# Fλ のランキングを計算する関数
def calculate_F_lambda_global_rank(df):
    global_ranks = []
    for column in df.columns:
        speed_data = df[column].values.flatten()
        eigenvalues, eigenvectors = eigh(L_normalized, eigvals_only=False)

        sorted_indices = np.argsort(eigenvalues)
        sorted_eigenvectors = eigenvectors[:, sorted_indices]

        # 各固有値に対応するフーリエ係数 Fλ を計算
        F_lambda = np.zeros(sorted_eigenvectors.shape[1], dtype=complex)
        for i in range(sorted_eigenvectors.shape[1]):
            u_lambda_i = sorted_eigenvectors[:, i]
            F_lambda[i] = sum(speed_data * np.conj(u_lambda_i))

        # Fλ の絶対値を計算し、ランキングを決定
        abs_F_lambda = np.abs(F_lambda)
        rank = abs_F_lambda.argsort().argsort()  # 小さい順のランキングを取得
        global_ranks.append(rank + 1)  # 1位が最大値になるよう反転（1位が最小値にならないように）

    return global_ranks

# シートにデータを書き込む関数
def write_sheet_with_ranks(writer, sheet_name, time_data, speed_data, dens_data, flow_sum, global_ranks):
    max_len = len(time_data)
    output_data = {
        'Time': list(time_data),
        'speed': list(speed_data),
        'dens': list(dens_data),
        'Flow Sum 19 Points': list(flow_sum),
    }
    # ランキングを各列に書き込む
    for lambda_index in range(20):
        output_data[f'Rank of F_lambda{lambda_index+1}'] = [
            ranks[lambda_index] if len(ranks) > lambda_index else '' for ranks in global_ranks
        ]

    # DataFrameに変換して書き込み
    results_df = pd.DataFrame(output_data)
    results_df.to_excel(writer, index=False, sheet_name=sheet_name)

# 各データフレームに対して処理を実行し、結果をシートに分けて保存
with pd.ExcelWriter(output_file, mode='w') as writer:
    for i, (df, sheet_name) in enumerate(zip(
        [df3, df4, df5, df6, df7, df8, df9, df10], 
        ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5', 'Sheet6', 'Sheet7', 'Sheet8']
    )):
        # 各時刻のFλランキングを計算
        global_ranks = calculate_F_lambda_global_rank(df)

        # Flow Sum 19 Points を計算
        flow_sum = (
            pd.Series(speed_data)
            .rolling(window=0, min_periods=0, center=True)
            .sum()
        )

        # シートにデータを書き込み
        write_sheet_with_ranks(writer, sheet_name, time_data, speed_data, dens_data, flow_sum, global_ranks)

print(f"{output_file} に固有値のグローバルランキングを含むデータを保存しました。")



print(f"{output_file}' が保存されました。")