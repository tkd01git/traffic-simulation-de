import pandas as pd
import openpyxl
import re
import os

def column_to_letter(column_index):
    """Excelの列番号をアルファベットに変換"""
    letter = ""
    while column_index >= 0:
        letter = chr(column_index % 26 + ord('A')) + letter
        column_index = column_index // 26 - 1
    return letter

# 最終編集先Excelファイル
excel_file_name = "F3_1000.xlsx"

# prodata.xlsxを読み込む
workbook = openpyxl.load_workbook("prodata1000.xlsx")
sheet = workbook["Sheet1"]  # Sheet1を対象とする

detector_col_index = None

# ヘッダー行（最初の行）をチェックして'列'のインデックスを取得
for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == "detector1":
        detector_col_index = col[0].column - 1  # 0ベースに変換
        break

if detector_col_index is not None:
    # 列番号をアルファベットに変換
    detector_col_letter = column_to_letter(detector_col_index)
    # 2つ前の列番号を計算
    if detector_col_index >= 2:
        two_columns_before_letter = column_to_letter(detector_col_index - 2)
    else:
        print("2つ前の列は存在しません。")
else:
    print("Detector1が見つかりません。")
import pandas as pd
import numpy as np
from scipy.linalg import eigh

file_path = "C://Users//YuheiTakada//OneDrive//デスクトップ//traffic-simulation-de//prodata1000.xlsx"
start_col = 'A'
end_col = two_columns_before_letter
sheet_name1 = 'Sheet1'
sheet_name2 = 'Sheet2'

# Sheet1とSheet2のデータの取得
df1 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None)
df2 = pd.read_excel(file_path, sheet_name=sheet_name2, usecols=f'{start_col}:{end_col}', header=None)

# 新しいデータフレームの作成
df3 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=1, nrows=20)  # 2~31行目
df4 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=4, nrows=20)  # 5~34行目
df5 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=7, nrows=20)  # 8~37行目
df6 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=10, nrows=20) # 11~40行目
df7 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=13, nrows=20)  # 5~34行目
df8 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=16, nrows=20)  # 5~34行目
df9 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=19, nrows=20)  # 5~34行目
df10 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=22, nrows=20)  # 5~34行目


time_data = df1.iloc[0].values.flatten()
flow_data = df2.iloc[21].values.flatten() / 3600  # 車両数データの取得
dens_data = df2.iloc[99].values.flatten()

# Laplacianの作成
n = 20
A = np.zeros((n, n))
for i in range(n - 1):
    A[i, i + 1] = 1
    A[i + 1, i] = 1
L = np.diag(np.sum(A, axis=1)) - A
L_normalized = np.dot(np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1)))), 
                      np.dot(L, np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1))))))

# F(λ_2) を計算する関数
def calculate_F_lambda_2(df):
    results = []
    for column in df.columns:
        speed_data = df[column].values.flatten()
        eigenvalues, eigenvectors = eigh(L_normalized, eigvals_only=False)

        sorted_indices = np.argsort(eigenvalues)
        sorted_eigenvectors = eigenvectors[:, sorted_indices]

        F_lambda = np.zeros(sorted_eigenvectors.shape[1], dtype=complex)
        for i in range(sorted_eigenvectors.shape[1]):
            u_lambda_i = sorted_eigenvectors[:, i]
            F_lambda[i] = sum(speed_data * np.conj(u_lambda_i))

        # Fλ_2 (インデックス1) の実部を取得
        results.append(F_lambda[2].real)
    return results

# Fλ_2 の計算
frambda2_df3 = calculate_F_lambda_2(df3)
frambda2_df4 = calculate_F_lambda_2(df4)
frambda2_df5 = calculate_F_lambda_2(df5)
frambda2_df6 = calculate_F_lambda_2(df6)
frambda2_df7 = calculate_F_lambda_2(df7)
frambda2_df8 = calculate_F_lambda_2(df8)
frambda2_df9 = calculate_F_lambda_2(df9)
frambda2_df10 = calculate_F_lambda_2(df10)

# 出力用のDataFrameを作成
output_data = {
    'Time': list(time_data),
    'flow': list(flow_data),
    'dens': list(dens_data),
    'Flow Sum 19 Points': [''] * len(time_data),  # 空白のまま作成
    'Franbda2_of_df3': frambda2_df3,
    'Franbda2_of_df4': frambda2_df4,
    'Franbda2_of_df5': frambda2_df5,
    'Franbda2_of_df6': frambda2_df6,
    'Franbda2_of_df7': frambda2_df7,
    'Franbda2_of_df8': frambda2_df8,
    'Franbda2_of_df9': frambda2_df9,
    'Franbda2_of_df10': frambda2_df10,
}

# 長さを調整してDataFrameに変換
max_len = max(len(time_data), len(frambda2_df3), len(frambda2_df4), len(frambda2_df5), len(frambda2_df6))
for key in output_data.keys():
    if len(output_data[key]) < max_len:
        output_data[key].extend([''] * (max_len - len(output_data[key])))

results_df = pd.DataFrame(output_data)

# Excelファイルにデータを書き込む
output_file = excel_file_name
with pd.ExcelWriter(output_file, mode='w') as writer:
    results_df.to_excel(writer, index=False, sheet_name='Results')

# Flow Sum 19 Points の計算を行い、Excelファイルを更新
results_df['Flow Sum 19 Points'] = results_df['flow'].rolling(window=19, min_periods=1, center=True).sum()

# Excelファイルに再度書き込む
with pd.ExcelWriter(output_file, mode='w') as writer:
    results_df.to_excel(writer, index=False, sheet_name='Results')

print(f"{output_file}' が保存されました。")