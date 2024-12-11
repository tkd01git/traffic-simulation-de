import pandas as pd
import numpy as np
from scipy.linalg import eigh
import openpyxl


def column_to_letter(column_index):
    """Excelの列番号をアルファベットに変換"""
    letter = ""
    while column_index >= 0:
        letter = chr(column_index % 26 + ord('A')) + letter
        column_index = column_index // 26 - 1
    return letter

# ファイルパス
file_path = "C://Users//YuheiTakada//OneDrive//デスクトップ//traffic-simulation-de//prodata800-15.xlsx"
output_file = "result15.xlsx"

# Excelファイルを読み込む
workbook = openpyxl.load_workbook(file_path)
sheet = workbook["Sheet1"]

detector_col_index = None

# ヘッダー行（最初の行）をチェックして'列'のインデックスを取得
for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == "detector1":
        detector_col_index = col[0].column - 1  # 0ベースに変換
        break

if detector_col_index is not None:
    detector_col_letter = column_to_letter(detector_col_index)
    if detector_col_index >= 2:
        two_columns_before_letter = column_to_letter(detector_col_index - 2)
    else:
        raise ValueError("2つ前の列は存在しません。")
else:
    raise ValueError("Detector1が見つかりません。")

# Sheet1とSheet2のデータの取得
start_col = 'A'
end_col = two_columns_before_letter
sheet_name1 = 'Sheet1'
sheet_name2 = 'Sheet2'

df1 = pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None)
df2 = pd.read_excel(file_path, sheet_name=sheet_name2, usecols=f'{start_col}:{end_col}', header=None)

# 新しいデータフレームを作成
data_frames = [
    pd.read_excel(file_path, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=i, nrows=20)
    for i in range(1, 25, 3)
]

time_data = df1.iloc[0].values.flatten()
speed_data = df2.iloc[44].values.flatten()
dens_data = df2.iloc[99].values.flatten()

# Laplacianの作成
n = 20
A = np.zeros((n, n))
for i in range(n - 1):
    A[i, i + 1] = 1
    A[i + 1, i] = 1
L = np.diag(np.sum(A, axis=1)) - A
L_normalized = np.dot(
    np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1)))),
    np.dot(L, np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1)))))
)

def calculate_F_lambda_real_parts(df):
    real_parts_list = []
    for column in df.columns:
        speed_data = df[column].values.flatten()
        eigenvalues, eigenvectors = eigh(L_normalized, eigvals_only=False)

        sorted_indices = np.argsort(eigenvalues)
        sorted_eigenvectors = eigenvectors[:, sorted_indices]

        F_lambda = np.zeros(sorted_eigenvectors.shape[1], dtype=complex)
        for i in range(sorted_eigenvectors.shape[1]):
            u_lambda_i = sorted_eigenvectors[:, i]
            F_lambda[i] = sum(speed_data * np.conj(u_lambda_i))

        real_parts = F_lambda.real
        real_parts_list.append(real_parts)
    return real_parts_list

def write_sheet_with_ranks(writer, sheet_name, time_data, speed_data, dens_data, flow_sum, real_parts_list):
    output_data = {
        'Time': list(time_data),
        'speed': list(speed_data),
        'Density': list(dens_data),
        'traffic volume': list(flow_sum),
    }
    for lambda_index in range(20):
        output_data[f'F_lambda{lambda_index+1}'] = [
            real_parts[lambda_index] if len(real_parts) > lambda_index else '' for real_parts in real_parts_list
        ]

    results_df = pd.DataFrame(output_data)
    results_df.to_excel(writer, index=False, sheet_name=sheet_name)

with pd.ExcelWriter(output_file, mode='w') as writer:
    for i, (df, sheet_name) in enumerate(zip(
        data_frames, 
        [f'Sheet{i+1}' for i in range(len(data_frames))]
    )):
        real_parts_list = calculate_F_lambda_real_parts(df)
        flow_sum = (
            pd.Series(speed_data)
            .rolling(window=0, min_periods=0)
            .sum()
        )
        write_sheet_with_ranks(writer, sheet_name, time_data, speed_data, dens_data, flow_sum, real_parts_list)

print(f"{output_file} が保存されました。")
