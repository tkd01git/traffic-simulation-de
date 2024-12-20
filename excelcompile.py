import pandas as pd
import openpyxl
import re
import os

# ファイルパス
file_path = "C://Users//Ytakada//Downloads//data (17).xlsx"

def process_detector_data(file_path, speed_column_index):
    # Excelファイルからすべてのデータを読み込む
    df = pd.read_excel(file_path, sheet_name="sheet1", engine='openpyxl')
    
    # 指定した列がデータの範囲を超える場合は空のDataFrameを返す
    if speed_column_index >= df.shape[1]:
        return pd.DataFrame()

    # 各カラムからデータを取得
    time_data = df.iloc[:, 0]
    data = df.iloc[:, speed_column_index]
    comments = df.iloc[:, 0]

    detector_data = {}
    detector_pattern = re.compile(r"detector(\d+)", re.IGNORECASE)

    detector_data['0'] = pd.DataFrame({
        'Time': time_data.iloc[:].values,  # 全行を取得
        'Data': data.iloc[:].values
    })

    current_detector = None
    current_time = []
    current_data = []

    for idx, comment in enumerate(comments):
        if isinstance(comment, str) and re.search(detector_pattern, comment):
            match = detector_pattern.search(comment)
            if match:
                detector_number = match.group(1)

                if current_detector is not None and len(current_time) > 0:
                    detector_data[current_detector] = pd.DataFrame({
                        'Time': current_time,
                        'Data': current_data
                    })

                current_detector = detector_number
                current_time = []
                current_data = []
        
        elif pd.notna(time_data.iloc[idx]) and pd.notna(data.iloc[idx]):
            current_time.append(time_data.iloc[idx])
            current_data.append(data.iloc[idx])

    if current_detector is not None and len(current_time) > 0:
        detector_data[current_detector] = pd.DataFrame({
            'Time': current_time,
            'Data': current_data
        })

    if not detector_data:
        return pd.DataFrame()

    combined_data = pd.DataFrame()

    for detector, data in detector_data.items():
        if combined_data.empty:
            combined_data['Time'] = data['Time']
        
        combined_data[f'Detector_{detector}'] = data['Data']
    
    combined_data.replace("--", 0, inplace=True)
    combined_data = combined_data.transpose()
    combined_data.reset_index(drop=True, inplace=True)

    return combined_data

def column_to_letter(column_index):
    """Excelの列番号をアルファベットに変換"""
    letter = ""
    while column_index >= 0:
        letter = chr(column_index % 26 + ord('A')) + letter
        column_index = column_index // 26 - 1
    return letter


# 速度データの処理
speed_combined_data = process_detector_data(file_path, 2)

# 流量データの処理
flow_combined_data = process_detector_data(file_path, 1)

# エクセルにデータを書き込む
with pd.ExcelWriter("prodata.xlsx", engine='openpyxl') as writer:
    if not speed_combined_data.empty:
        speed_combined_data.to_excel(writer, sheet_name="Sheet1", index=False, header=False)
    if not flow_combined_data.empty:
        flow_combined_data.to_excel(writer, sheet_name="Sheet2", index=False, header=False)

# Sheet2のデータを処理
sheet2_data = pd.read_excel("prodata.xlsx", sheet_name="Sheet2", header=None, engine='openpyxl')

# Sheet1のデータを再度読み込む
sheet1_data = pd.read_excel("prodata.xlsx", sheet_name="Sheet1", header=None, engine='openpyxl')

# 必要な行数を確認し、不足している場合は行を追加
required_rows = 60
if sheet2_data.shape[0] < required_rows:
    additional_rows = required_rows - sheet2_data.shape[0]
    sheet2_data = pd.concat([sheet2_data, pd.DataFrame([[None] * sheet2_data.shape[1]] * additional_rows, columns=sheet2_data.columns)], ignore_index=True)

# 各列の2〜42行目の値を掛け合わせてその列の総和を43行目に記入
for col in range(sheet1_data.shape[1]):
    product_sum = (sheet1_data.iloc[1:42, col] * sheet2_data.iloc[1:42, col]).sum()
    sheet2_data.iloc[42, col] = product_sum

# 各列の2〜42行目の総和を計算して44行目に記入
for col in range(sheet2_data.shape[1]):
    col_sum = sheet2_data.iloc[1:42, col].sum()
    sheet2_data.iloc[43, col] = col_sum

# 各列の43行目を44行目で割り、小数第2位で四捨五入して45行目に記入
for col in range(sheet2_data.shape[1]):
    value_23 = sheet2_data.iloc[42, col]
    value_24 = sheet2_data.iloc[43, col]
    
    if pd.notna(value_23) and pd.notna(value_24) and value_24 != 0:
        result = round(value_23 / value_24, 2)
    else:
        result = None
    
    sheet2_data.iloc[44, col] = result
    
# 各列の2~21行目で、sheet1_dataとsheet2_dataのセルを掛け合わせて、その総和を46行目に書き込む
for col in range(sheet1_data.shape[1]):
    product_sum = (sheet1_data.iloc[1:22, col] * sheet2_data.iloc[1:22, col]).sum()
    sheet2_data.iloc[45, col] = product_sum

# 各列のsheet2_dataの2~21行目の総和を47行目に書き込む
for col in range(sheet2_data.shape[1]):
    col_sum = sheet2_data.iloc[1:22, col].sum()
    sheet2_data.iloc[46, col] = col_sum

# 各列の46行目を47行目で割り、小数第2位で四捨五入して48行目に書き込む
for col in range(sheet2_data.shape[1]):
    value_25 = sheet2_data.iloc[45, col]
    value_26 = sheet2_data.iloc[46, col]
    
    if pd.notna(value_25) and pd.notna(value_26) and value_26 != 0:
        result = round(value_25 / value_26, 2)
    else:
        result = None
    
    sheet2_data.iloc[47, col] = result
    
# 各列の22~41行目で、sheet1_dataとsheet2_dataのセルを掛け合わせて、その総和を49行目に書き込む
for col in range(sheet1_data.shape[1]):
    product_sum = (sheet1_data.iloc[22:42, col] * sheet2_data.iloc[22:42, col]).sum()
    sheet2_data.iloc[48, col] = product_sum

# 各列のsheet2_dataの22~41行目の総和を50行目に書き込む
for col in range(sheet2_data.shape[1]):
    col_sum = sheet2_data.iloc[22:41, col].sum()
    sheet2_data.iloc[49, col] = col_sum

# 各列の49行目を50行目で割り、小数第2位で四捨五入して51行目に書き込む
for col in range(sheet2_data.shape[1]):
    value_28 = sheet2_data.iloc[48, col]
    value_29 = sheet2_data.iloc[49, col]
    
    if pd.notna(value_28) and pd.notna(value_29) and value_29 != 0:
        result = round(value_28 / value_29, 2)
    else:
        result = None
    sheet2_data.iloc[50, col] = result


# 追加：4列目のデータをSheet2の52行目に書き込む
df_additional_data = pd.read_excel(file_path, sheet_name="sheet1", engine='openpyxl')
if df_additional_data.shape[1] >= 4:
    sheet2_data.iloc[51, :] = df_additional_data.iloc[:, 3].values[:sheet2_data.shape[1]]


    
    
# 最後に全てのデータを書き込む
with pd.ExcelWriter("prodata.xlsx", engine='openpyxl', mode='w') as writer:
    sheet1_data.to_excel(writer, sheet_name="Sheet1", index=False, header=False)
    sheet2_data.to_excel(writer, sheet_name="Sheet2", index=False, header=False)

print("Data has been successfully written to 'prodata.xlsx'.")


# prodata.xlsxを読み込む
workbook = openpyxl.load_workbook("prodata.xlsx")
sheet = workbook["Sheet1"]  # Sheet1を対象とする

detector_col_index = None

# ヘッダー行（最初の行）をチェックして'列'のインデックスを取得
for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == "detector1":
        detector_col_index = col[0].column - 1  # 0ベースに変換
        break

if detector_col_index is not None:
    # 列番号をアルファベットに変換
    detector_col_letter = column_to_letter(detector_col_index)
    # 2つ前の列番号を計算
    if detector_col_index >= 2:
        two_columns_before_letter = column_to_letter(detector_col_index - 2)
        print(f"Detector1の列番号: {detector_col_letter}")
        print(f"2つ前の列番号: {two_columns_before_letter}")
    else:
        print("2つ前の列は存在しません。")
else:
    print("Detector1が見つかりません。")